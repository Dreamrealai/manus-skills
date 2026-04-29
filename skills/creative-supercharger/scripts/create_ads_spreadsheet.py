#!/usr/bin/env python3
"""
Create the Example Ads Database spreadsheet template.
Usage: python3 create_ads_spreadsheet.py [BRAND_NAME] [OUTPUT_PATH]
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_ads_spreadsheet(brand_name: str, output_path: str):
    wb = openpyxl.Workbook()

    # --- Summary Tab ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    subheader_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    header_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    accent_fill = PatternFill(start_color="C8A951", end_color="C8A951", fill_type="solid")

    ws_summary["A1"] = f"{brand_name} — Example Ads Database"
    ws_summary["A1"].font = header_font
    ws_summary["A1"].fill = header_fill
    ws_summary.merge_cells("A1:D1")

    summary_rows = [
        ("Total Ads Cataloged (Brand)", ""),
        ("Total Ads Cataloged (Competitor)", ""),
        ("Videos Downloaded", ""),
        ("", ""),
        ("Breakdown by Platform", "Count"),
        ("Meta (Facebook/Instagram)", ""),
        ("TikTok", ""),
        ("YouTube", ""),
        ("Google Ads", ""),
        ("LinkedIn", ""),
        ("Other", ""),
        ("", ""),
        ("Breakdown by Format", "Count"),
        ("Video", ""),
        ("Carousel", ""),
        ("Single Image", ""),
        ("Collection", ""),
        ("", ""),
        ("Top 3 Messaging Themes", ""),
        ("1.", ""),
        ("2.", ""),
        ("3.", ""),
        ("", ""),
        ("Top 3 Creative Styles", ""),
        ("1.", ""),
        ("2.", ""),
        ("3.", ""),
        ("", ""),
        ("Notable Patterns", ""),
        ("", ""),
    ]

    for i, (col_a, col_b) in enumerate(summary_rows, start=3):
        ws_summary[f"A{i}"] = col_a
        ws_summary[f"B{i}"] = col_b
        ws_summary[f"A{i}"].font = body_font
        ws_summary[f"B{i}"].font = body_font

    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 30

    # --- Ads Database Tab ---
    ws_ads = wb.create_sheet("Ads Database")

    columns = [
        ("Ad_ID", 8),
        ("Brand", 20),
        ("Platform", 25),
        ("Format", 15),
        ("Status", 12),
        ("First_Seen_Date", 16),
        ("Ad_Copy_Headline", 35),
        ("Ad_Copy_Body", 45),
        ("CTA_Button", 15),
        ("Core_Message", 40),
        ("Hook_Type", 20),
        ("Production_Style", 18),
        ("Landing_Page_URL", 40),
        ("Thumbnail_Saved", 15),
        ("Video_Downloaded", 16),
        ("Notes", 40),
    ]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws_ads.cell(row=1, column=col_idx, value=col_name)
        cell.font = subheader_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws_ads.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Add 30 empty rows with borders
    for row_idx in range(2, 32):
        for col_idx in range(1, len(columns) + 1):
            cell = ws_ads.cell(row=row_idx, column=col_idx, value="")
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Freeze top row
    ws_ads.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Spreadsheet created: {output_path}")


if __name__ == "__main__":
    brand = sys.argv[1] if len(sys.argv) > 1 else "BRAND"
    output = sys.argv[2] if len(sys.argv) > 2 else "Example_Ads.xlsx"
    create_ads_spreadsheet(brand, output)
