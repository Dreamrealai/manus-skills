#!/usr/bin/env python3
"""
brandinsights/scripts/source_excel_builder.py

Extracts all Markdown citations from insights.md and builds sources.xlsx.
Supports both [Source Name](URL) inline links and (Source Name, Month YYYY) parenthetical citations.

Usage:
  python3 source_excel_builder.py --input insights.md --output sources.xlsx

Output columns:
  Topic | Claim | Source Name | URL | Stream | Source Date | Date Accessed | Verification Status | Hallucination Risk | Notes
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pandas and openpyxl are required. Run: sudo pip3 install pandas openpyxl")
    sys.exit(1)


# Hallucination risk keywords that trigger HIGH or CRITICAL risk
HIGH_RISK_KEYWORDS = [
    "hhs", "fda", "usda", "government", "official", "poison", "toxic",
    "color of the year", "award", "record", "all-time", "unprecedented",
    "first ever", "never before", "guaranteed", "proven"
]

CRITICAL_RISK_KEYWORDS = [
    "hhs calls", "fda says", "government says", "officially declared",
    "poison", "banned", "illegal", "criminal", "fraud"
]


def score_hallucination_risk(claim: str, source_name: str, url: str) -> tuple[str, str]:
    """
    Score hallucination risk for a claim.
    Returns (risk_level, notes) where risk_level is Low/Medium/High/Critical.
    """
    claim_lower = claim.lower()
    source_lower = source_name.lower()
    notes = []

    # Critical risk: government quotes without official source
    if any(kw in claim_lower for kw in CRITICAL_RISK_KEYWORDS):
        return "Critical", "Contains government/official quote that may be misattributed. Verify before use."

    # Critical risk: no URL and unusual claim
    if (not url or url == "N/A") and any(kw in claim_lower for kw in HIGH_RISK_KEYWORDS):
        return "Critical", "No URL provided and claim contains high-risk language. Verify before use."

    # High risk: no URL
    if not url or url == "N/A" or url == "":
        notes.append("No URL provided")
        return "High", "No source URL available. Single-source claim. Verify independently."

    # High risk: unusual claims
    if any(kw in claim_lower for kw in HIGH_RISK_KEYWORDS):
        notes.append("Contains high-risk language")
        return "High", f"Claim contains language that requires verification: {', '.join([kw for kw in HIGH_RISK_KEYWORDS if kw in claim_lower])}"

    # Medium risk: social media source
    if any(x in url.lower() for x in ["reddit.com", "twitter.com", "x.com", "instagram.com", "tiktok.com"]):
        return "Medium", "Social media source. Verify with primary source."

    # Medium risk: no date in citation
    if "n/a" in source_lower or "unknown" in source_lower:
        return "Medium", "Source date unknown. Verify recency."

    # Low risk: primary source with URL
    return "Low", "Primary source with URL. Independently verifiable."


def extract_citations(text: str) -> list[dict]:
    """Extract all citations from markdown text.
    
    Supports:
    1. Inline links: [Source Name](URL)
    2. Reference-style links: [1]: https://...
    3. Parenthetical citations: (Source Name, Month YYYY) or (Source Name, YYYY)
    """
    rows = []
    current_topic = "General"
    current_claim = ""

    for line in text.splitlines():
        # Track section headers as topics
        if line.startswith("## "):
            current_topic = line.lstrip("# ").strip()
        elif line.startswith("### "):
            current_topic = line.lstrip("# ").strip()

        # 1. Extract inline citations: [Source Name](URL)
        citations = re.findall(r'\[([^\]]+)\]\((https?://[^\)]+)\)', line)
        if citations:
            claim_text = re.sub(r'\[([^\]]+)\]\((https?://[^\)]+)\)', '', line).strip()
            claim_text = claim_text.lstrip("- *").strip()
            for source_name, url in citations:
                risk_level, risk_notes = score_hallucination_risk(claim_text, source_name, url)
                rows.append({
                    "Topic": current_topic,
                    "Claim": claim_text[:300] if claim_text else "See source",
                    "Source Name": source_name,
                    "URL": url,
                    "Stream": _infer_stream(url, source_name),
                    "Source Date": _extract_date_from_source(source_name),
                    "Date Accessed": datetime.now().strftime("%Y-%m-%d"),
                    "Verification Status": "Unverified",
                    "Hallucination Risk": risk_level,
                    "Notes": risk_notes,
                })

        # 2. Extract reference-style links: [1]: https://...
        ref_match = re.match(r'^\[(\d+)\]:\s*(https?://\S+)\s*"?([^"]*)"?', line)
        if ref_match:
            ref_num, url, title = ref_match.groups()
            risk_level, risk_notes = score_hallucination_risk(title or "", title or "", url)
            rows.append({
                "Topic": current_topic,
                "Claim": f"Reference [{ref_num}]",
                "Source Name": title or f"Source {ref_num}",
                "URL": url,
                "Stream": _infer_stream(url, title),
                "Source Date": "N/A",
                "Date Accessed": datetime.now().strftime("%Y-%m-%d"),
                "Verification Status": "Unverified",
                "Hallucination Risk": risk_level,
                "Notes": risk_notes,
            })

        # 3. Extract parenthetical citations: (Source Name, Month YYYY) or (Source Name, YYYY)
        paren_citations = re.findall(
            r'\(([A-Za-z][^,\(\)]{2,50}),\s*((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4}|\d{4}|Q[1-4]\s+\d{4})\)',
            line
        )
        if paren_citations:
            claim_text = re.sub(
                r'\(([A-Za-z][^,\(\)]{2,50}),\s*((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4}|\d{4}|Q[1-4]\s+\d{4})\)',
                '', line
            ).strip()
            claim_text = claim_text.lstrip("- *").strip()
            for source_name, source_date in paren_citations:
                source_name = source_name.strip()
                # Skip if this looks like a non-citation (e.g., "e.g.", "i.e.", "vs.")
                if len(source_name) < 3 or source_name.lower() in ["e.g", "i.e", "vs", "etc"]:
                    continue
                risk_level, risk_notes = score_hallucination_risk(claim_text, source_name, "")
                rows.append({
                    "Topic": current_topic,
                    "Claim": claim_text[:300] if claim_text else "See source",
                    "Source Name": source_name,
                    "URL": "N/A (parenthetical citation — add URL)",
                    "Stream": _infer_stream("", source_name),
                    "Source Date": source_date.strip(),
                    "Date Accessed": datetime.now().strftime("%Y-%m-%d"),
                    "Verification Status": "Unverified",
                    "Hallucination Risk": risk_level,
                    "Notes": f"Parenthetical citation. {risk_notes}",
                })

    return rows


def _extract_date_from_source(source_name: str) -> str:
    """Try to extract a date from a source name."""
    # Match patterns like "Q4 2025", "2025", "January 2026", "Jan 2026"
    date_match = re.search(
        r'(Q[1-4]\s+\d{4}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4}|\d{4})',
        source_name
    )
    if date_match:
        return date_match.group(1)
    return "N/A"


def _infer_stream(url: str, source_name: str) -> str:
    """Infer which research stream produced this citation."""
    url_lower = url.lower() if url else ""
    source_lower = source_name.lower() if source_name else ""
    if any(x in url_lower for x in ["reddit.com", "twitter.com", "x.com", "instagram.com", "linkedin.com", "youtube.com", "tiktok.com"]):
        return "Social Mining"
    if any(x in source_lower for x in ["gemini", "google"]):
        return "Gemini Deep Research"
    if any(x in source_lower for x in ["chatgpt", "openai"]):
        return "ChatGPT Deep Research"
    if any(x in source_lower for x in ["claude", "anthropic"]):
        return "Claude Deep Research"
    return "Manus Research"


def build_excel(rows: list[dict], output_path: str) -> None:
    if not rows:
        print("WARNING: No citations found. Generating empty sources.xlsx.")
        rows = [{
            "Topic": "N/A", "Claim": "No citations extracted", "Source Name": "N/A",
            "URL": "N/A", "Stream": "N/A", "Source Date": "N/A", "Date Accessed": "N/A",
            "Verification Status": "N/A", "Hallucination Risk": "N/A", "Notes": "N/A"
        }]

    df = pd.DataFrame(rows).drop_duplicates(subset=["Source Name", "Claim"])
    df = df.sort_values(["Hallucination Risk", "Topic", "Stream"])

    # Define column order
    columns = ["Topic", "Claim", "Source Name", "URL", "Stream", "Source Date", "Date Accessed", "Verification Status", "Hallucination Risk", "Notes"]
    df = df.reindex(columns=columns, fill_value="N/A")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sources")
        ws = writer.sheets["Sources"]

        # Style header row
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Color-code Hallucination Risk column
        risk_col_idx = columns.index("Hallucination Risk") + 1
        risk_colors = {
            "Low": "C6EFCE",      # Green
            "Medium": "FFEB9C",   # Yellow
            "High": "FFCC99",     # Orange
            "Critical": "FFC7CE", # Red
        }
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=risk_col_idx)
            risk_level = str(cell.value)
            if risk_level in risk_colors:
                cell.fill = PatternFill(start_color=risk_colors[risk_level], end_color=risk_colors[risk_level], fill_type="solid")
                cell.font = Font(bold=True)

        # Auto-size columns
        col_widths = {
            "Topic": 25, "Claim": 60, "Source Name": 30, "URL": 50, "Stream": 20,
            "Source Date": 15, "Date Accessed": 15, "Verification Status": 18,
            "Hallucination Risk": 18, "Notes": 40
        }
        for i, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 20)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add hyperlinks to URL column
        url_col_idx = columns.index("URL") + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=url_col_idx)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    # Print risk summary
    risk_counts = df["Hallucination Risk"].value_counts()
    print(f"\n✅ sources.xlsx written to {output_path} ({len(df)} unique citations)")
    print("\n📊 Hallucination Risk Summary:")
    for risk in ["Critical", "High", "Medium", "Low"]:
        count = risk_counts.get(risk, 0)
        if count > 0:
            print(f"   {risk}: {count} claims")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build sources.xlsx from insights.md citations")
    parser.add_argument("--input", required=True, help="Path to insights.md")
    parser.add_argument("--output", required=True, help="Path to output sources.xlsx")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    text = input_path.read_text(encoding="utf-8")
    rows = extract_citations(text)
    print(f"Extracted {len(rows)} citations from {args.input}")
    build_excel(rows, args.output)


if __name__ == "__main__":
    main()
